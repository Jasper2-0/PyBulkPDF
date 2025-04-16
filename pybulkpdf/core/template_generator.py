# -*- coding: utf-8 -*-
"""
Handles the generation of template files (Excel and Field Info TXT)
based on an input PDF form template. Uses class-based design and custom exceptions.
"""

import os
import sys
import logging
from typing import List, Tuple, Dict, Any, Optional, Set

# --- Third-Party Libraries ---
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet # For type hinting
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

# --- Relative Imports ---
from .. import config
from ..utils.validation import check_file_exists
# Import custom exceptions
from ..exceptions import PDFReadError, ExcelWriteError, FileOperationError
from .pdf_analyzer import PDFAnalyzer

# --- Constants removed, using config module ---

class TemplateGenerator:
    """
    Generates template files (Excel, Field Info TXT) for a given PDF.

    Uses PDFAnalyzer to get field information. Raises custom exceptions on failure.
    """
    def __init__(self, template_pdf_path: str, output_dir: str):
        """
        Initializes the TemplateGenerator.

        Args:
            template_pdf_path (str): Path to the input PDF form template.
            output_dir (str): Directory to save the generated template files.

        Raises:
            FileOperationError: If the output directory is invalid.
            PDFReadError: If the PDFAnalyzer fails to initialize (propagated).
        """
        if not os.path.isdir(output_dir):
             msg = f"Output directory '{output_dir}' does not exist or is not a directory."
             logging.error(msg)
             raise FileOperationError(msg) # Use custom exception

        self.template_pdf_path: str = template_pdf_path
        self.output_dir: str = output_dir
        self.base_filename: str = os.path.splitext(os.path.basename(template_pdf_path))[0]

        # Create analyzer instance - this might raise PDFReadError
        self.analyzer: PDFAnalyzer = PDFAnalyzer(template_pdf_path)
        # If PDFAnalyzer init succeeded but found no fields, it's handled later

    def _generate_excel_template(self) -> str:
        """
        Generates the Excel template file (.xlsx) with headers and table formatting.

        Returns:
            The full path to the generated Excel file.

        Raises:
            ExcelWriteError: If creating or saving the Excel file fails.
            ConfigurationError: If PDF fields couldn't be loaded (should not happen if init succeeded).
        """
        field_names_set = self.analyzer.get_field_names()
        # This check is slightly redundant if PDFReadError is handled by caller,
        # but provides safety within the method.
        if field_names_set is None:
             # This state implies PDF reading failed during __init__
             msg = "Cannot generate Excel template: PDF fields were not loaded."
             logging.error(msg)
             # Or raise a different exception type? Let's stick to ExcelWriteError context
             raise ExcelWriteError(msg)

        field_names = sorted(list(field_names_set))
        xlsx_filename = f"{self.base_filename}{config.TEMPLATE_SUFFIX}"
        xlsx_filepath = os.path.join(self.output_dir, xlsx_filename)
        xlsx_headers = field_names + [config.OUTPUT_FILENAME_COL]

        try:
            wb = Workbook()
            ws = wb.active
            if ws is None: # Should not happen
                raise ExcelWriteError("Could not get active worksheet from new workbook.")
            ws.title = config.DEFAULT_SHEET_NAME
            ws.append(xlsx_headers)

            if xlsx_headers:
                last_col_letter = get_column_letter(len(xlsx_headers))
                table_range = f"A1:{last_col_letter}{2}"
                tab = Table(displayName=config.DEFAULT_TABLE_NAME, ref=table_range)
                style = TableStyleInfo(name=config.EXCEL_TABLE_STYLE, showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                logging.debug(f"Formatted Excel range ({table_range}) as Table ('{config.DEFAULT_TABLE_NAME}').")

                for col_idx, header in enumerate(xlsx_headers, 1):
                     column_letter = get_column_letter(col_idx)
                     adjusted_width = len(str(header)) + 5
                     ws.column_dimensions[column_letter].width = adjusted_width
            else:
                logging.warning("Skipping Excel Table creation as no headers were generated.")

            wb.save(xlsx_filepath) # This operation can fail
            logging.info(f"Generated Excel template: {xlsx_filepath}")
            return xlsx_filepath

        except (OSError, InvalidFileException, Exception) as e:
            msg = f"Failed to create or save Excel template '{xlsx_filepath}'"
            logging.error(f"{msg}: {e}")
            # Raise custom exception, wrapping original
            raise ExcelWriteError(msg, original_exception=e)

    def _generate_field_info_file(self) -> Optional[str]:
        """
        Generates the field information text file (_field_info.txt).

        Returns:
            The full path to the generated text file, or None if no info was generated.

        Raises:
            FileOperationError: If writing the text file fails.
        """
        field_info_list = self.analyzer.analyze_field_types()
        if not field_info_list:
            logging.info("No specific non-text field information found to generate.")
            return None # Not an error, just nothing to write

        txt_filename = f"{self.base_filename}{config.FIELD_INFO_SUFFIX}"
        txt_filepath = os.path.join(self.output_dir, txt_filename)

        try:
            with open(txt_filepath, 'w', encoding='utf-8') as txtfile:
                txtfile.write("Information about expected values for non-text PDF fields:\n")
                txtfile.write(f"Note: Checkbox/Radio Button values are often '{config.PDF_VALUE_CHECKBOX_ON}' or the specific value shown.\n")
                txtfile.write("If unsure, test with a single row first.\n")
                txtfile.write("=========================================================\n\n")
                for line in field_info_list:
                    txtfile.write(line + "\n")
            logging.info(f"Generated field info file: {txt_filepath}")
            return txt_filepath
        except (OSError, Exception) as e:
            # Only raise exception on actual write errors
            msg = f"Failed to write field info file '{txt_filepath}'"
            logging.error(f"{msg}: {e}")
            raise FileOperationError(msg, original_exception=e) # Raise custom exception

    def generate_files(self) -> None:
        """
        Orchestrates the generation of both Excel and Field Info files.

        Raises:
            ExcelWriteError: If Excel generation fails.
            FileOperationError: If Field Info file writing fails.
            (PDFReadError is handled by the caller via __init__)
        """
        logging.info(f"Starting template file generation process for {self.template_pdf_path}")

        # Generate Excel - raises ExcelWriteError on failure
        self._generate_excel_template()

        # Generate Field Info - raises FileOperationError on failure
        # This runs only if Excel generation succeeded
        self._generate_field_info_file()

        logging.info("Template generation process completed.")
        # No return value needed, success is indicated by lack of exceptions


# --- Public Function (Interface for CLI) ---

def generate_template_files(template_pdf_path: str, output_dir: str) -> None:
    """
    Public interface function to generate template files.
    Instantiates and uses the TemplateGenerator class, handles exceptions.

    Args:
        template_pdf_path: Path to the input PDF form template.
        output_dir: Directory where the generated template files will be saved.
    """
    try:
        # Check input file before potentially lengthy operations
        check_file_exists(template_pdf_path) # Raises FileOperationError

        # Instantiate generator (might raise PDFReadError or FileOperationError)
        generator = TemplateGenerator(template_pdf_path, output_dir)

        # Run generation (might raise ExcelWriteError or FileOperationError)
        generator.generate_files()

    # Catch specific custom exceptions first
    except (FileOperationError, PDFReadError, ExcelWriteError) as e:
        # Log the specific error captured by our custom exceptions
        logging.critical(f"Template generation failed: {e}")
        sys.exit(1) # Exit with error status
    # Catch any other unexpected exceptions
    except Exception as e:
        logging.critical(f"An unexpected error occurred during template generation: {e}", exc_info=True)
        sys.exit(1)

