# -*- coding: utf-8 -*-
"""
Handles the core logic for filling PDF forms using data from an Excel file.
Uses class-based design and custom exceptions.
"""

import os
import sys
import logging
from typing import List, Tuple, Dict, Any, Optional, Set, Iterable, Union

# --- Third-Party Libraries ---
from pypdf import PdfReader, PdfWriter
from pypdf import errors as pypdf_errors
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from tqdm import tqdm

# --- Relative Imports ---
from .. import config
from ..utils.validation import check_file_exists
# Import custom exceptions
from ..exceptions import (
    PyBulkPDFError, PDFReadError, PDFWriteError, ExcelReadError,
    ConfigurationError, FileOperationError
)

# --- Type Aliases ---
RowData = Tuple[Any, ...]
ExcelHeaders = List[str]
ProcessingResult = Tuple[bool, str] # (Success Status, Reason/Message)
FailedRowInfo = Tuple[int, str]

# --- FormFiller Class ---
class FormFiller:
    """
    Fills PDF forms based on Excel data using an object-oriented approach.
    Raises specific exceptions on critical setup failures. Handles row-level
    errors during processing.
    """
    def __init__(self, template_path: str, data_path: str, output_dir: str, overwrite: bool = False):
        """
        Initializes the FormFiller instance.

        Args:
            template_path (str): Path to the input PDF form template.
            data_path (str): Path to the input Excel (.xlsx) data file.
            output_dir (str): Directory where the filled PDF files will be saved.
            overwrite (bool): If True, allows overwriting existing files.

        Raises:
            FileOperationError: If input paths/directories are invalid.
        """
        self.template_path: str = template_path
        self.data_path: str = data_path
        self.output_dir: str = output_dir
        self.overwrite: bool = overwrite

        # Initialize state variables
        self.pdf_field_names: Optional[Set[str]] = None
        self.xlsx_headers: Optional[ExcelHeaders] = None
        self.sheet: Optional[Worksheet] = None
        self.total_rows: int = 0
        self.common_fields: Optional[Set[str]] = None

        self.row_count: int = 0
        self.success_count: int = 0
        self.failed_rows: List[FailedRowInfo] = []

        # Basic validation - raise FileOperationError if invalid
        if not os.path.isfile(template_path):
            raise FileOperationError(f"Template PDF not found or is not a file: {template_path}")
        if not os.path.isfile(data_path):
            raise FileOperationError(f"Data file not found or is not a file: {data_path}")
        if not os.path.isdir(output_dir):
            raise FileOperationError(f"Output directory not found or is not a directory: {output_dir}")

        logging.info(f"FormFiller initialized for template: {self.template_path}")
        logging.info(f"Data source: {self.data_path}")
        logging.info(f"Output directory: {self.output_dir} (Overwrite: {self.overwrite})")

    def _read_pdf_template_fields(self) -> None:
        """
        Reads PDF template fields and stores them in self.pdf_field_names.

        Raises:
            PDFReadError: If reading or parsing the PDF fails.
        """
        try:
            reader = PdfReader(self.template_path)
            fields = reader.get_fields()
            if not fields:
                # Treat as a configuration issue if template has no fields
                msg = f"No fillable fields found in template PDF: {self.template_path}"
                logging.error(msg)
                raise ConfigurationError(msg) # Changed from PDFReadError
            self.pdf_field_names = set(fields.keys())
            logging.info(f"Template PDF fields found: {len(self.pdf_field_names)}")
        except pypdf_errors.PdfReadError as e:
            msg = f"Error reading PDF structure from '{self.template_path}'"
            logging.error(f"{msg}: {e}")
            raise PDFReadError(msg, original_exception=e)
        except Exception as e:
            msg = f"Unexpected error reading PDF template fields '{self.template_path}'"
            logging.error(f"{msg}: {e}")
            raise PDFReadError(msg, original_exception=e)

    def _read_excel_data(self) -> None:
        """
        Opens Excel, reads headers/sheet, stores them.

        Raises:
            ExcelReadError: If reading the Excel file fails.
        """
        try:
            workbook: Workbook = openpyxl.load_workbook(self.data_path, data_only=True)
            sheet = workbook.active
            if sheet is None:
                raise ExcelReadError(f"Could not get active sheet from Excel file: {self.data_path}")
            self.sheet = sheet

            header_iter = self.sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            xlsx_headers_raw = next(header_iter, None)
            if xlsx_headers_raw is None:
                raise ExcelReadError(f"Excel file has no header row: {self.data_path}")

            self.xlsx_headers = [str(h).strip() if h is not None else '' for h in xlsx_headers_raw]
            if not self.xlsx_headers or all(h == '' for h in self.xlsx_headers):
                raise ExcelReadError(f"Excel file has no valid headers in the first row: {self.data_path}")

            self.total_rows = self.sheet.max_row - 1 if self.sheet.max_row > 1 else 0
            if self.total_rows <= 0:
                logging.warning(f"No data rows found in Excel file: {self.data_path}")

        except (InvalidFileException, FileNotFoundError, StopIteration, Exception) as e:
            msg = f"Error reading Excel file {self.data_path}"
            logging.error(f"{msg}: {e}")
            raise ExcelReadError(msg, original_exception=e)

    def _validate_headers_and_map_fields(self) -> None:
        """
        Validates headers, compares with PDF fields, stores common fields.

        Raises:
            ConfigurationError: If required column is missing or no common fields found.
        """
        if self.pdf_field_names is None or self.xlsx_headers is None:
             # Should not happen if called after successful setup
             raise ConfigurationError("Cannot validate headers: PDF fields or Excel headers not loaded.")

        xlsx_header_set = set(self.xlsx_headers)
        if config.OUTPUT_FILENAME_COL not in xlsx_header_set:
            msg = f"Required column '{config.OUTPUT_FILENAME_COL}' not found in Excel file headers."
            logging.error(msg)
            raise ConfigurationError(msg)

        xlsx_data_fields = xlsx_header_set - {config.OUTPUT_FILENAME_COL}
        pdf_only_fields = self.pdf_field_names - xlsx_data_fields
        xlsx_only_fields = xlsx_data_fields - self.pdf_field_names

        if pdf_only_fields: logging.warning(f"PDF fields not found in Excel headers: {', '.join(sorted(list(pdf_only_fields)))}")
        if xlsx_only_fields: logging.warning(f"Excel headers not found in PDF fields: {', '.join(sorted(list(xlsx_only_fields)))}")

        self.common_fields = self.pdf_field_names.intersection(xlsx_data_fields)
        if not self.common_fields:
            msg = "No common fields found between PDF template and Excel headers. Cannot proceed."
            logging.error(msg)
            raise ConfigurationError(msg)

        logging.info(f"Fields to be filled: {len(self.common_fields)} ({', '.join(sorted(list(self.common_fields)))})")

    def _prepare_fill_data(self, row_dict: Dict[str, Any]) -> Dict[str, Union[str, bool]]:
        """Prepares the data dictionary for pypdf."""
        fill_data = {}
        if self.common_fields is None: return {} # Should not happen
        for field in self.common_fields:
            value = row_dict.get(field)
            if isinstance(value, bool):
                fill_data[field] = config.PDF_VALUE_CHECKBOX_ON if value else config.PDF_VALUE_CHECKBOX_OFF
            elif value is None:
                fill_data[field] = ''
            else:
                fill_data[field] = str(value)
        return fill_data # type: ignore

    def _fill_single_pdf(self, fill_data: Dict[str, Any], output_filepath: str) -> None:
        """
        Fills a single PDF file using the provided data.

        Raises:
            PDFReadError: If cloning the template fails.
            PDFWriteError: If writing the filled PDF fails.
        """
        writer = None
        try:
            # Clone template - can raise PDFReadError
            writer = PdfWriter(clone_from=self.template_path)

            # Update fields
            for page_index, page in enumerate(writer.pages):
                try:
                    writer.update_page_form_field_values(page, fields=fill_data)
                except KeyError: pass
                except Exception as page_update_error: # Log non-critical page update errors
                    logging.warning(f"Page {page_index+1} update issue for {os.path.basename(output_filepath)}: {page_update_error}")

            # Clean up appearances flag
            if writer._root_object and "/AcroForm" in writer._root_object:
                acro_form = writer._root_object["/AcroForm"]
                if hasattr(acro_form, 'get') and "/NeedAppearances" in acro_form:
                    acro_form.pop("/NeedAppearances")

            # Write file - can raise various OS/pypdf errors
            with open(output_filepath, "wb") as output_stream:
                writer.write(output_stream)

        except pypdf_errors.PdfReadError as e: # Error during cloning
            msg = f"Template read error during cloning for {os.path.basename(output_filepath)}"
            # Log error before raising
            logging.error(f"{msg}: {e}")
            raise PDFReadError(msg, original_exception=e)
        except (FileNotFoundError, PermissionError, Exception) as e: # Errors during write
            msg = f"PDF write error for {os.path.basename(output_filepath)}"
            logging.error(f"{msg}: {e}")
            raise PDFWriteError(msg, original_exception=e)
        finally:
            writer = None # Ensure cleanup

    def _process_single_row(self, row_num: int, row_values: RowData) -> Optional[ProcessingResult]:
        """
        Processes a single row of Excel data. Handles expected row-level
        errors and returns status tuple. Catches PDF write errors.

        Returns:
            Tuple (bool, str): (Success Status, Reason/Message)
            None: If the row was skipped (e.g., completely empty).
        """
        if all(v is None for v in row_values): return None
        if self.xlsx_headers is None or self.common_fields is None:
            # This indicates a programming error if setup succeeded
            return False, "Internal error: Headers/common fields not available."

        row_dict = {header: value for header, value in zip(self.xlsx_headers, row_values)}
        output_filename_raw = row_dict.get(config.OUTPUT_FILENAME_COL)
        current_output_filename = str(output_filename_raw).strip() if output_filename_raw is not None else ''

        # --- Handle expected row-level skips/failures ---
        if not current_output_filename:
            return False, f"'{config.OUTPUT_FILENAME_COL}' is empty"
        if not current_output_filename.lower().endswith('.pdf'):
            current_output_filename += '.pdf'
        output_filepath = os.path.join(self.output_dir, current_output_filename)
        if not self.overwrite and os.path.exists(output_filepath):
            return False, f"Output file exists: {current_output_filename} (use --overwrite)"

        # --- Prepare data and fill PDF (catching errors) ---
        try:
            fill_data = self._prepare_fill_data(row_dict)
            # Call PDF filling, which now raises exceptions on failure
            self._fill_single_pdf(fill_data, output_filepath)
            # If _fill_single_pdf completes without exception, it succeeded
            return True, f"Successfully generated {current_output_filename}"

        except (PDFReadError, PDFWriteError) as e:
            # Catch specific PDF errors from _fill_single_pdf
            # Logged already in _fill_single_pdf, just return failure status
            return False, f"PDF processing error: {e}"
        except Exception as e:
            # Catch unexpected errors during data prep or other issues
            logging.error(f"Unexpected error processing row {row_num} data: {e}", exc_info=True)
            return False, f"Unexpected error: {e}"

    def _generate_summary_report(self) -> None:
        """Logs the final summary."""
        # (Implementation unchanged)
        logging.info("-" * 40)
        logging.info("Form Filling Process Summary:")
        logging.info(f"Total data rows processed: {self.row_count}")
        logging.info(f"Successfully generated PDFs: {self.success_count}")
        failures = len(self.failed_rows)
        if failures > 0:
            logging.warning(f"Rows with failures/skips: {failures}")
            for row_num, reason in self.failed_rows: logging.warning(f"  - Row {row_num}: {reason}")
        elif self.row_count > 0: logging.info("All processed rows generated PDFs successfully.")
        else: logging.info("No data rows were processed.")
        logging.info("-" * 40)

    def run(self) -> None:
        """
        Executes the entire form filling process. Handles setup exceptions.

        Raises:
            PyBulkPDFError: If a critical setup step fails (PDFReadError,
                           ExcelReadError, ConfigurationError).
        """
        # 1. Perform setup steps - exceptions will halt execution if they occur
        self._read_pdf_template_fields() # Raises PDFReadError or ConfigurationError
        self._read_excel_data()          # Raises ExcelReadError
        self._validate_headers_and_map_fields() # Raises ConfigurationError

        # Ensure sheet is available (should be if setup succeeded)
        if self.sheet is None:
             raise PyBulkPDFError("Internal error: Excel sheet not loaded after setup.")

        # 2. Process Rows
        row_iterator = self.sheet.iter_rows(min_row=2, values_only=True)
        progress_bar = tqdm(row_iterator, total=self.total_rows, desc="Filling PDFs", unit="row",
                            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}{postfix}]",
                            ncols=100, disable=self.total_rows <= 0)

        for row_index, row_values in enumerate(progress_bar):
            row_num = row_index + 2
            result = self._process_single_row(row_num, row_values) # Handles its own exceptions internally now

            if result is None: continue # Skip empty row
            self.row_count += 1
            success, message = result

            if success:
                self.success_count += 1
                filename = next((s for s in message.split() if s.lower().endswith('.pdf')), f"row {row_num}")
                progress_bar.set_postfix_str(f"Processed {filename}", refresh=True)
            else:
                self.failed_rows.append((row_num, message))
                logging.warning(f"Skipping row {row_num}: {message}") # Log skip reason
                progress_bar.set_postfix_str(f"Failed row {row_num}", refresh=True)

        progress_bar.close()

        # 3. Generate Summary
        self._generate_summary_report()
        # Success is indicated by reaching the end without setup exceptions


# --- Public Function (Interface for CLI) ---
def fill_pdf_forms(template_pdf_path: str, data_file_path: str, output_dir: str, overwrite: bool = False) -> None:
    """
    Public interface function to fill PDF forms. Handles exceptions from FormFiller.
    """
    try:
        # Check input files before potentially lengthy operations
        check_file_exists(template_pdf_path) # Raises FileOperationError
        check_file_exists(data_file_path)   # Raises FileOperationError

        # Instantiate and run (can raise FileOperationError, PDFReadError, ExcelReadError, ConfigurationError)
        filler = FormFiller(template_pdf_path, data_file_path, output_dir, overwrite)
        filler.run()

    # Catch specific custom exceptions from setup/run
    except (FileOperationError, PDFReadError, ExcelReadError, ConfigurationError) as e:
        logging.critical(f"Form filling failed: {e}")
        sys.exit(1) # Exit with error status
    # Catch any other unexpected exceptions
    except Exception as e:
        logging.critical(f"An unexpected error occurred during form filling: {e}", exc_info=True)
        sys.exit(1)

