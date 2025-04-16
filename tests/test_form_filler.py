# -*- coding: utf-8 -*-
"""
Unit tests for the pybulkpdf.core.form_filler module using pytest.
"""

import pytest
import sys
import os
from unittest.mock import MagicMock, call, ANY

# --- Modules/Classes to test ---
from pybulkpdf.core import form_filler
from pybulkpdf.core.form_filler import FormFiller
from pybulkpdf.exceptions import (
    PyBulkPDFError, PDFReadError, PDFWriteError, ExcelReadError,
    ConfigurationError, FileOperationError
)
from pybulkpdf import config
from pybulkpdf.utils.validation import check_file_exists # For patching

# --- Third-party library components to mock ---
from pypdf import errors as pypdf_errors
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = MagicMock(); Workbook = MagicMock; Worksheet = MagicMock; InvalidFileException = Exception
try: from pypdf import PdfReader, PdfWriter
except ImportError: PdfReader = MagicMock; PdfWriter = MagicMock
try: from tqdm import tqdm
except ImportError: tqdm = lambda x, **kwargs: x

# --- Mock Data ---
MOCK_TEMPLATE_PATH = "input/template.pdf"
MOCK_DATA_PATH = "input/data.xlsx"
MOCK_OUTPUT_DIR = "output/filled"
MOCK_PDF_FIELDS = {"Name", "Date", "Approved"}
MOCK_EXCEL_HEADERS = ["Name", "Date", "Approved", config.OUTPUT_FILENAME_COL]
MOCK_COMMON_FIELDS = {"Name", "Date", "Approved"}
MOCK_ROW_1_VALS = ("Alice", "2025-04-16", True, "alice_report.pdf")
MOCK_ROW_2_VALS = ("Bob", "2025-04-17", False, "bob_report")
MOCK_ROW_3_VALS = ("Charlie", None, None, "")
MOCK_ROW_4_VALS = ("David", "2025-04-18", "Invalid", "david_report.pdf")
MOCK_EXCEL_DATA_ITER = [ MOCK_ROW_1_VALS, MOCK_ROW_2_VALS, MOCK_ROW_3_VALS, MOCK_ROW_4_VALS, (None, None, None, None)]


# --- Fixtures ---

@pytest.fixture
def mock_dependencies(mocker):
    """Fixture to mock external dependencies like os, openpyxl, pypdf, tqdm."""
    mock_tqdm_instance = MagicMock()
    mock_tqdm_instance.__iter__.return_value = iter(MOCK_EXCEL_DATA_ITER)
    mock_tqdm_instance.set_postfix_str = MagicMock()
    mock_tqdm_instance.close = MagicMock()

    mocks = {
        'isfile': mocker.patch('os.path.isfile', return_value=True),
        'isdir': mocker.patch('os.path.isdir', return_value=True),
        'exists': mocker.patch('os.path.exists', return_value=False),
        'join': mocker.patch('os.path.join', lambda *args: "/".join(args)),
        'PdfReader': mocker.patch('pybulkpdf.core.form_filler.PdfReader'),
        'PdfWriter': mocker.patch('pybulkpdf.core.form_filler.PdfWriter'),
        'load_workbook': mocker.patch('pybulkpdf.core.form_filler.openpyxl.load_workbook'),
        'open': mocker.patch('builtins.open', mocker.mock_open()),
        'tqdm': mocker.patch('pybulkpdf.core.form_filler.tqdm', return_value=mock_tqdm_instance),
        'check_file_exists': mocker.patch('pybulkpdf.core.form_filler.check_file_exists'),
        'sys_exit': mocker.patch('sys.exit'),
        'tqdm_instance': mock_tqdm_instance
    }
    mock_reader_instance = MagicMock()
    mock_reader_instance.get_fields.return_value = {f: {} for f in MOCK_PDF_FIELDS}
    mocks['PdfReader'].return_value = mock_reader_instance
    mock_writer_instance = MagicMock()
    mock_writer_instance.write = MagicMock()
    mock_writer_instance.update_page_form_field_values = MagicMock()
    mock_writer_instance._root_object = {"/AcroForm": {"/NeedAppearances": True}}
    mock_writer_instance.pages = [MagicMock()]
    mocks['PdfWriter'].return_value = mock_writer_instance
    mock_sheet = MagicMock(spec=Worksheet)
    mock_sheet.max_row = len(MOCK_EXCEL_DATA_ITER) + 1
    # Adjust side_effect to provide fresh iterators each time iter_rows is called in tests
    mocks['mock_sheet'] = mock_sheet # Store sheet mock itself
    def iter_rows_side_effect(*args, **kwargs):
        if kwargs.get('max_row') == 1:
            return iter([MOCK_EXCEL_HEADERS])
        else:
            return iter(MOCK_EXCEL_DATA_ITER)
    mock_sheet.iter_rows.side_effect = iter_rows_side_effect

    mock_workbook = MagicMock(spec=Workbook)
    mock_workbook.active = mock_sheet
    mocks['load_workbook'].return_value = mock_workbook
    return mocks

@pytest.fixture
def form_filler_instance(mock_dependencies):
    """Fixture to create a FormFiller instance with mocked dependencies."""
    try:
        instance = FormFiller(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR, overwrite=False)
        # Reset iter_rows side effect for the instance's sheet object if needed,
        # although the fixture should provide a fresh one each time.
        instance.pdf_field_names = MOCK_PDF_FIELDS
        instance.xlsx_headers = MOCK_EXCEL_HEADERS
        instance.sheet = mock_dependencies['mock_sheet'] # Use the stored sheet mock
        instance.total_rows = len(MOCK_EXCEL_DATA_ITER) # Correct calculation
        instance.common_fields = MOCK_COMMON_FIELDS
        return instance
    except Exception as e:
        pytest.fail(f"FormFiller initialization failed in fixture: {e}")


# --- Tests for FormFiller Initialization ---
# (No changes needed here)
def test_form_filler_init_success(mock_dependencies):
    try:
        filler = FormFiller(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
        assert filler.template_path == MOCK_TEMPLATE_PATH
    except Exception as e: pytest.fail(f"Init raised {e}")
def test_form_filler_init_bad_template_path(mock_dependencies):
    mock_dependencies['isfile'].side_effect = lambda p: False if p == MOCK_TEMPLATE_PATH else True
    with pytest.raises(FileOperationError): FormFiller(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
def test_form_filler_init_bad_data_path(mock_dependencies):
    mock_dependencies['isfile'].side_effect = lambda p: False if p == MOCK_DATA_PATH else True
    with pytest.raises(FileOperationError): FormFiller(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
def test_form_filler_init_bad_output_dir(mock_dependencies):
    mock_dependencies['isdir'].return_value = False
    with pytest.raises(FileOperationError): FormFiller(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)

# --- Tests for Setup Methods (implicitly via run / public function) ---
# (No changes needed here)
def test_run_handles_pdf_read_error(mocker, mock_dependencies):
    mocker.patch.object(FormFiller, '_read_pdf_template_fields', side_effect=PDFReadError("PDF Read Fail"))
    mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    mock_dependencies['sys_exit'].assert_called_once_with(1)
def test_run_handles_excel_read_error(mocker, mock_dependencies):
    mocker.patch.object(FormFiller, '_read_excel_data', side_effect=ExcelReadError("Excel Read Fail"))
    mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    mock_dependencies['sys_exit'].assert_called_once_with(1)
def test_run_handles_config_error(mocker, mock_dependencies):
    mocker.patch.object(FormFiller, '_validate_headers_and_map_fields', side_effect=ConfigurationError("Config Fail"))
    mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    mock_dependencies['sys_exit'].assert_called_once_with(1)

# --- Tests for _prepare_fill_data ---
# (No changes needed here)
@pytest.mark.parametrize("input_val, expected_val", [
    ("Test String", "Test String"), (123, "123"), (45.6, "45.6"),
    (True, config.PDF_VALUE_CHECKBOX_ON), (False, config.PDF_VALUE_CHECKBOX_OFF), (None, ""),
])
def test_prepare_fill_data_conversions(form_filler_instance, input_val, expected_val):
    form_filler_instance.common_fields = {"TestData"}
    row_dict = {"TestData": input_val}
    result = form_filler_instance._prepare_fill_data(row_dict)
    assert result == {"TestData": expected_val}

# --- Tests for _fill_single_pdf ---
# (No changes needed here)
def test_fill_single_pdf_success(mocker, mock_dependencies, form_filler_instance):
    mock_writer = mock_dependencies['PdfWriter'].return_value; mock_open = mock_dependencies['open']
    fill_data = {"Name": "Test"}; output_path = f"{MOCK_OUTPUT_DIR}/output.pdf"
    form_filler_instance._fill_single_pdf(fill_data, output_path)
    mock_dependencies['PdfWriter'].assert_called_once_with(clone_from=MOCK_TEMPLATE_PATH)
    assert mock_writer.update_page_form_field_values.call_count == 1
    mock_open.assert_called_once_with(output_path, "wb")
    mock_writer.write.assert_called_once_with(mock_open.return_value.__enter__.return_value)
def test_fill_single_pdf_clone_fails(mocker, mock_dependencies, form_filler_instance):
    mocker.patch('pybulkpdf.core.form_filler.PdfWriter', side_effect=pypdf_errors.PdfReadError("Clone fail"))
    fill_data = {"Name": "Test"}; output_path = f"{MOCK_OUTPUT_DIR}/output.pdf"
    with pytest.raises(PDFReadError): form_filler_instance._fill_single_pdf(fill_data, output_path)
def test_fill_single_pdf_write_fails(mocker, mock_dependencies, form_filler_instance):
    mock_writer = mock_dependencies['PdfWriter'].return_value; mock_writer.write.side_effect = OSError("Disk full")
    mock_open = mock_dependencies['open']; fill_data = {"Name": "Test"}; output_path = f"{MOCK_OUTPUT_DIR}/output.pdf"
    with pytest.raises(PDFWriteError): form_filler_instance._fill_single_pdf(fill_data, output_path)
    mock_open.assert_called_once_with(output_path, "wb")
    mock_writer.write.assert_called_once_with(mock_open.return_value.__enter__.return_value)

# --- Tests for _process_single_row ---
# (No changes needed here)
def test_process_single_row_success(mocker, form_filler_instance):
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data', return_value={"Name": "Alice"})
    mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 2; row_vals = MOCK_ROW_1_VALS
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is True; assert "Successfully generated alice_report.pdf" in result[1]
    mock_prepare.assert_called_once()
    mock_fill.assert_called_once_with({"Name": "Alice"}, f"{MOCK_OUTPUT_DIR}/alice_report.pdf")
def test_process_single_row_empty_filename(mocker, form_filler_instance):
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data'); mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 4; row_vals = MOCK_ROW_3_VALS
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is False; assert f"'{config.OUTPUT_FILENAME_COL}' is empty" in result[1]
    mock_prepare.assert_not_called(); mock_fill.assert_not_called()
def test_process_single_row_file_exists_no_overwrite(mocker, form_filler_instance):
    mocker.patch('os.path.exists', return_value=True)
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data'); mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 2; row_vals = MOCK_ROW_1_VALS; form_filler_instance.overwrite = False
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is False; assert "Output file exists: alice_report.pdf" in result[1]
    mock_prepare.assert_not_called(); mock_fill.assert_not_called()
def test_process_single_row_file_exists_with_overwrite(mocker, form_filler_instance):
    mocker.patch('os.path.exists', return_value=True)
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data', return_value={"Name": "Alice"})
    mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 2; row_vals = MOCK_ROW_1_VALS; form_filler_instance.overwrite = True
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is True
    mock_prepare.assert_called_once(); mock_fill.assert_called_once()
def test_process_single_row_prepare_fails(mocker, form_filler_instance):
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data', side_effect=TypeError("Bad data type"))
    mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 2; row_vals = MOCK_ROW_1_VALS
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is False; assert "Unexpected error: Bad data type" in result[1]
    mock_prepare.assert_called_once(); mock_fill.assert_not_called()
def test_process_single_row_fill_fails(mocker, form_filler_instance):
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data', return_value={"Name": "Alice"})
    mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf', side_effect=PDFWriteError("Disk full"))
    row_num = 2; row_vals = MOCK_ROW_1_VALS
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is False; assert "PDF processing error: Disk full" in result[1]
    mock_prepare.assert_called_once(); mock_fill.assert_called_once()
def test_process_single_row_adds_pdf_extension(mocker, form_filler_instance):
    mock_prepare = mocker.patch.object(form_filler_instance, '_prepare_fill_data', return_value={"Name": "Bob"})
    mock_fill = mocker.patch.object(form_filler_instance, '_fill_single_pdf')
    row_num = 3; row_vals = MOCK_ROW_2_VALS
    result = form_filler_instance._process_single_row(row_num, row_vals)
    assert result[0] is True
    expected_output_path = f"{MOCK_OUTPUT_DIR}/bob_report.pdf"
    mock_fill.assert_called_once_with({"Name": "Bob"}, expected_output_path)
    assert "Successfully generated bob_report.pdf" in result[1]

# --- Tests for run() Method ---

def test_run_success_all_rows(mocker, mock_dependencies, form_filler_instance):
    """Test run method completes successfully processing all rows."""
    # *** FIX: Use side_effect to simulate skipping the last empty row ***
    process_results = [(True, "Success")] * 4 + [None] # 4 successes, 1 skip
    mock_process = mocker.patch.object(FormFiller, '_process_single_row', side_effect=process_results)
    mock_summary = mocker.patch.object(FormFiller, '_generate_summary_report')
    # Mock setup methods called by run
    mocker.patch.object(form_filler_instance, '_read_pdf_template_fields')
    mocker.patch.object(form_filler_instance, '_read_excel_data')
    mocker.patch.object(form_filler_instance, '_validate_headers_and_map_fields')

    form_filler_instance.run()

    mock_tqdm_instance = mock_dependencies['tqdm_instance']
    # *** FIX: Assert based on corrected expectation (4 successes) ***
    assert mock_tqdm_instance.set_postfix_str.call_count == 4 # Called only for 4 successes
    assert mock_tqdm_instance.close.call_count == 1

    assert mock_process.call_count == len(MOCK_EXCEL_DATA_ITER) # Called for all 5 rows
    # *** FIX: Assert counts based on 4 processed rows ***
    assert form_filler_instance.row_count == 4 # Only non-skipped rows counted
    assert form_filler_instance.success_count == 4
    assert form_filler_instance.failed_rows == []
    mock_summary.assert_called_once()

# (test_run_success_mixed_results remains the same)
def test_run_success_mixed_results(mocker, mock_dependencies, form_filler_instance):
    process_results = [(True, "Success 1"), (False, "File exists"), (False, "Empty filename"), (True, "Success 4"), None]
    mock_process = mocker.patch.object(FormFiller, '_process_single_row', side_effect=process_results)
    mock_summary = mocker.patch.object(FormFiller, '_generate_summary_report')
    mocker.patch.object(form_filler_instance, '_read_pdf_template_fields')
    mocker.patch.object(form_filler_instance, '_read_excel_data')
    mocker.patch.object(form_filler_instance, '_validate_headers_and_map_fields')
    form_filler_instance.run()
    mock_tqdm_instance = mock_dependencies['tqdm_instance']
    assert mock_tqdm_instance.set_postfix_str.call_count == 4 # Called for 2 success, 2 fail
    assert mock_tqdm_instance.close.call_count == 1
    assert mock_process.call_count == len(MOCK_EXCEL_DATA_ITER)
    assert form_filler_instance.row_count == 4
    assert form_filler_instance.success_count == 2
    assert len(form_filler_instance.failed_rows) == 2
    assert (4, "Empty filename") in form_filler_instance.failed_rows
    assert (3, "File exists") in form_filler_instance.failed_rows
    mock_summary.assert_called_once()

# --- Tests for fill_pdf_forms (Public Function) ---
# (No changes needed here)
def test_public_fill_pdf_forms_success(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    mock_form_filler_cls = mocker.patch('pybulkpdf.core.form_filler.FormFiller')
    mock_sys_exit = mocker.patch('sys.exit')
    mock_filler_instance = MagicMock(); mock_filler_instance.run.return_value = None
    mock_form_filler_cls.return_value = mock_filler_instance
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    assert mock_check_file.call_count == 2
    mock_form_filler_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR, False)
    mock_filler_instance.run.assert_called_once(); mock_sys_exit.assert_not_called()
def test_public_fill_pdf_forms_check_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.form_filler.check_file_exists', side_effect=FileOperationError("Check fail"))
    mock_sys_exit = mocker.patch('sys.exit')
    mocker.patch('pybulkpdf.core.form_filler.FormFiller')
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    mock_check_file.assert_called_once(); mock_sys_exit.assert_called_once_with(1)
def test_public_fill_pdf_forms_init_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    mock_form_filler_cls = mocker.patch('pybulkpdf.core.form_filler.FormFiller', side_effect=ConfigurationError("Init fail"))
    mock_sys_exit = mocker.patch('sys.exit')
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    assert mock_check_file.call_count == 2
    mock_form_filler_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR, False)
    mock_sys_exit.assert_called_once_with(1)
def test_public_fill_pdf_forms_run_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.form_filler.check_file_exists')
    mock_form_filler_cls = mocker.patch('pybulkpdf.core.form_filler.FormFiller')
    mock_sys_exit = mocker.patch('sys.exit')
    mock_filler_instance = MagicMock(); mock_filler_instance.run.side_effect = ExcelReadError("Run setup fail")
    mock_form_filler_cls.return_value = mock_filler_instance
    form_filler.fill_pdf_forms(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR)
    assert mock_check_file.call_count == 2
    mock_form_filler_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_DATA_PATH, MOCK_OUTPUT_DIR, False)
    mock_filler_instance.run.assert_called_once(); mock_sys_exit.assert_called_once_with(1)

