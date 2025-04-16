# -*- coding: utf-8 -*-
"""
Unit tests for the pybulkpdf.core.template_generator module using pytest.
"""

import pytest
import sys
import os
from unittest.mock import MagicMock, call

# --- Modules/Classes to test ---
from pybulkpdf.core import template_generator
from pybulkpdf.core.template_generator import TemplateGenerator
from pybulkpdf.exceptions import PDFReadError, ExcelWriteError, FileOperationError, ConfigurationError
from pybulkpdf import config
from pybulkpdf.core.pdf_analyzer import PDFAnalyzer
from pybulkpdf.utils.validation import check_file_exists

# --- Third-party Mocks ---
try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    Workbook = MagicMock; Worksheet = MagicMock; Table = MagicMock; TableStyleInfo = MagicMock; get_column_letter = MagicMock(return_value='XFD'); InvalidFileException = Exception


# --- Mock Data ---
MOCK_TEMPLATE_PATH = "input/template.pdf"
MOCK_OUTPUT_DIR = "output/templates"
MOCK_BASE_FILENAME = "template"
MOCK_FIELD_NAMES = {"FieldA", "FieldB", "Checkbox1"}
MOCK_ANALYSIS_LIST = ["Field 'Checkbox1' (Button): Expected values..."]

# --- Fixtures ---

@pytest.fixture
def mock_pdf_analyzer(mocker):
    """Fixture to create a mock PDFAnalyzer instance."""
    mock_analyzer = MagicMock(spec=PDFAnalyzer)
    mock_analyzer.get_field_names.return_value = MOCK_FIELD_NAMES
    mock_analyzer.analyze_field_types.return_value = MOCK_ANALYSIS_LIST
    mocker.patch('pybulkpdf.core.template_generator.PDFAnalyzer', return_value=mock_analyzer)
    return mock_analyzer

@pytest.fixture
def mock_workbook(mocker):
    """Fixture to create mock openpyxl Workbook and Worksheet."""
    mock_ws = MagicMock(spec=Worksheet)
    # *** FIX: Configure column_dimensions mock ***
    mock_dimension = MagicMock() # Mock object returned by column_dimensions['X']
    mock_ws.column_dimensions = MagicMock() # Mock the column_dimensions attribute itself
    # Make it behave like a dict returning our mock dimension object when accessed
    mock_ws.column_dimensions.__getitem__.return_value = mock_dimension

    mock_wb = MagicMock(spec=Workbook)
    mock_wb.active = mock_ws
    mock_wb.save = MagicMock()
    mocker.patch('pybulkpdf.core.template_generator.Workbook', return_value=mock_wb)
    mocker.patch('pybulkpdf.core.template_generator.get_column_letter', return_value='X')
    mocker.patch('pybulkpdf.core.template_generator.TableStyleInfo')
    mocker.patch('pybulkpdf.core.template_generator.Table')
    return mock_wb, mock_ws

@pytest.fixture
def mock_os_path(mocker):
    """Fixture to mock os.path functions."""
    mocker.patch('os.path.isdir', return_value=True)
    mocker.patch('os.path.splitext', return_value=(MOCK_BASE_FILENAME, ".pdf"))
    mocker.patch('os.path.basename', return_value=f"{MOCK_BASE_FILENAME}.pdf")
    mocker.patch('os.path.join', lambda *args: "/".join(args))


# --- Tests for TemplateGenerator Class ---
# (Initialization tests remain the same)
def test_template_generator_init_success(mock_os_path, mock_pdf_analyzer):
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    assert generator.template_pdf_path == MOCK_TEMPLATE_PATH
    assert generator.output_dir == MOCK_OUTPUT_DIR
    assert generator.base_filename == MOCK_BASE_FILENAME
    assert generator.analyzer == mock_pdf_analyzer
    template_generator.PDFAnalyzer.assert_called_once_with(MOCK_TEMPLATE_PATH)

def test_template_generator_init_output_dir_not_found(mocker):
    mocker.patch('os.path.isdir', return_value=False)
    mocker.patch('os.path.splitext', return_value=(MOCK_BASE_FILENAME, ".pdf"))
    mocker.patch('os.path.basename', return_value=f"{MOCK_BASE_FILENAME}.pdf")
    with pytest.raises(FileOperationError, match="Output directory .* does not exist"):
        TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)

def test_template_generator_init_pdf_analyzer_fails(mocker, mock_os_path):
    mocker.patch('pybulkpdf.core.template_generator.PDFAnalyzer', side_effect=PDFReadError("Mock PDF Read Fail"))
    with pytest.raises(PDFReadError, match="Mock PDF Read Fail"):
        TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)

# --- Tests for _generate_excel_template ---

def test_generate_excel_template_success(mocker, mock_os_path, mock_pdf_analyzer, mock_workbook):
    """Test successful Excel template generation."""
    mock_wb, mock_ws = mock_workbook
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    expected_path = f"{MOCK_OUTPUT_DIR}/{MOCK_BASE_FILENAME}{config.TEMPLATE_SUFFIX}"
    result_path = generator._generate_excel_template() # Should not raise error now

    assert result_path == expected_path
    expected_headers = sorted(list(MOCK_FIELD_NAMES)) + [config.OUTPUT_FILENAME_COL]
    mock_ws.append.assert_called_once_with(expected_headers)
    assert mock_ws.add_table.call_count == 1
    # *** Check column_dimensions was accessed ***
    assert mock_ws.column_dimensions.__getitem__.call_count == len(expected_headers)
    mock_wb.save.assert_called_once_with(expected_path)

def test_generate_excel_template_no_fields(mocker, mock_os_path, mock_pdf_analyzer, mock_workbook):
    """Test Excel generation when PDF has no fields."""
    mock_wb, mock_ws = mock_workbook
    mock_pdf_analyzer.get_field_names.return_value = set()
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    expected_path = f"{MOCK_OUTPUT_DIR}/{MOCK_BASE_FILENAME}{config.TEMPLATE_SUFFIX}"
    result_path = generator._generate_excel_template() # Should not raise error now

    assert result_path == expected_path
    mock_ws.append.assert_called_once_with([config.OUTPUT_FILENAME_COL])
    assert mock_ws.add_table.call_count == 1
    # *** Check column_dimensions was accessed ***
    assert mock_ws.column_dimensions.__getitem__.call_count == 1 # Only for output filename col
    mock_wb.save.assert_called_once_with(expected_path)

# (analyzer_failed test remains the same)
def test_generate_excel_template_analyzer_failed(mocker, mock_os_path, mock_pdf_analyzer):
    mock_pdf_analyzer.get_field_names.return_value = None
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    with pytest.raises(ExcelWriteError, match="PDF fields were not loaded"):
        generator._generate_excel_template()

def test_generate_excel_template_save_fails(mocker, mock_os_path, mock_pdf_analyzer, mock_workbook):
    """Test Excel generation raises ExcelWriteError if save fails."""
    mock_wb, mock_ws = mock_workbook
    mock_wb.save.side_effect = OSError("Disk full")
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    expected_path = f"{MOCK_OUTPUT_DIR}/{MOCK_BASE_FILENAME}{config.TEMPLATE_SUFFIX}"

    with pytest.raises(ExcelWriteError, match="Failed to create or save Excel template"):
        generator._generate_excel_template() # Will fail on save now

    # *** Assert save was called (it should be reached now) ***
    mock_wb.save.assert_called_once_with(expected_path)


# --- Tests for _generate_field_info_file ---
# (No changes needed here)
def test_generate_field_info_success(mocker, mock_os_path, mock_pdf_analyzer):
    mock_open = mocker.patch('builtins.open', mocker.mock_open())
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    expected_path = f"{MOCK_OUTPUT_DIR}/{MOCK_BASE_FILENAME}{config.FIELD_INFO_SUFFIX}"
    result_path = generator._generate_field_info_file()
    assert result_path == expected_path
    mock_pdf_analyzer.analyze_field_types.assert_called_once()
    mock_open.assert_called_once_with(expected_path, 'w', encoding='utf-8')
    handle = mock_open()
    handle.write.assert_any_call(MOCK_ANALYSIS_LIST[0] + "\n")

def test_generate_field_info_no_info(mocker, mock_os_path, mock_pdf_analyzer):
    mock_pdf_analyzer.analyze_field_types.return_value = []
    mock_open = mocker.patch('builtins.open', mocker.mock_open())
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    result_path = generator._generate_field_info_file()
    assert result_path is None
    mock_pdf_analyzer.analyze_field_types.assert_called_once()
    mock_open.assert_not_called()

def test_generate_field_info_write_fails(mocker, mock_os_path, mock_pdf_analyzer):
    mock_open = mocker.patch('builtins.open', mocker.mock_open())
    mock_open.side_effect = OSError("Permission denied")
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    expected_path = f"{MOCK_OUTPUT_DIR}/{MOCK_BASE_FILENAME}{config.FIELD_INFO_SUFFIX}"
    with pytest.raises(FileOperationError, match="Failed to write field info file"):
        generator._generate_field_info_file()
    mock_pdf_analyzer.analyze_field_types.assert_called_once()
    mock_open.assert_called_once_with(expected_path, 'w', encoding='utf-8')


# --- Tests for generate_files method ---
# (No changes needed here)
def test_generate_files_success(mocker, mock_os_path, mock_pdf_analyzer):
    mock_gen_excel = mocker.patch.object(TemplateGenerator, '_generate_excel_template', return_value="path/to/excel.xlsx")
    mock_gen_info = mocker.patch.object(TemplateGenerator, '_generate_field_info_file', return_value="path/to/info.txt")
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    generator.generate_files()
    mock_gen_excel.assert_called_once()
    mock_gen_info.assert_called_once()

def test_generate_files_excel_fails(mocker, mock_os_path, mock_pdf_analyzer):
    mock_gen_excel = mocker.patch.object(TemplateGenerator, '_generate_excel_template', side_effect=ExcelWriteError("Excel fail"))
    mock_gen_info = mocker.patch.object(TemplateGenerator, '_generate_field_info_file')
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    with pytest.raises(ExcelWriteError, match="Excel fail"):
        generator.generate_files()
    mock_gen_excel.assert_called_once()
    mock_gen_info.assert_not_called()

def test_generate_files_info_fails(mocker, mock_os_path, mock_pdf_analyzer):
    mock_gen_excel = mocker.patch.object(TemplateGenerator, '_generate_excel_template', return_value="path/to/excel.xlsx")
    mock_gen_info = mocker.patch.object(TemplateGenerator, '_generate_field_info_file', side_effect=FileOperationError("Info fail"))
    generator = TemplateGenerator(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    with pytest.raises(FileOperationError, match="Info fail"):
        generator.generate_files()
    mock_gen_excel.assert_called_once()
    mock_gen_info.assert_called_once()


# --- Tests for generate_template_files (Public Function) ---
# (No changes needed here)
def test_public_generate_template_files_success(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.template_generator.check_file_exists')
    mock_template_generator_cls = mocker.patch('pybulkpdf.core.template_generator.TemplateGenerator')
    mock_sys_exit = mocker.patch('sys.exit')
    mock_generator_instance = MagicMock()
    mock_generator_instance.generate_files.return_value = None
    mock_template_generator_cls.return_value = mock_generator_instance
    template_generator.generate_template_files(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_check_file.assert_called_once_with(MOCK_TEMPLATE_PATH)
    mock_template_generator_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_generator_instance.generate_files.assert_called_once()
    mock_sys_exit.assert_not_called()

def test_public_generate_template_files_check_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.template_generator.check_file_exists', side_effect=FileOperationError("File check fail"))
    mock_sys_exit = mocker.patch('sys.exit')
    mocker.patch('pybulkpdf.core.template_generator.TemplateGenerator')
    template_generator.generate_template_files(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_check_file.assert_called_once_with(MOCK_TEMPLATE_PATH)
    mock_sys_exit.assert_called_once_with(1)

def test_public_generate_template_files_init_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.template_generator.check_file_exists')
    mock_template_generator_cls = mocker.patch('pybulkpdf.core.template_generator.TemplateGenerator', side_effect=PDFReadError("Init fail"))
    mock_sys_exit = mocker.patch('sys.exit')
    template_generator.generate_template_files(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_check_file.assert_called_once_with(MOCK_TEMPLATE_PATH)
    mock_template_generator_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_sys_exit.assert_called_once_with(1)

def test_public_generate_template_files_generate_fails(mocker):
    mock_check_file = mocker.patch('pybulkpdf.core.template_generator.check_file_exists')
    mock_template_generator_cls = mocker.patch('pybulkpdf.core.template_generator.TemplateGenerator')
    mock_sys_exit = mocker.patch('sys.exit')
    mock_generator_instance = MagicMock()
    mock_generator_instance.generate_files.side_effect = ExcelWriteError("Generate fail")
    mock_template_generator_cls.return_value = mock_generator_instance
    template_generator.generate_template_files(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_check_file.assert_called_once_with(MOCK_TEMPLATE_PATH)
    mock_template_generator_cls.assert_called_once_with(MOCK_TEMPLATE_PATH, MOCK_OUTPUT_DIR)
    mock_generator_instance.generate_files.assert_called_once()
    mock_sys_exit.assert_called_once_with(1)
