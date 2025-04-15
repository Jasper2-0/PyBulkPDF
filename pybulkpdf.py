#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PyBulkPDF: A Command-Line Tool for PDF Form Mail Merge using Excel Data.

This script allows users to automate the process of filling PDF form fields
using data provided in an Excel (.xlsx) spreadsheet. It offers two main modes:

1.  `generate-template`: Analyzes a given PDF form template, identifies its
    fillable fields, and generates:
    a) An Excel (.xlsx) template file with headers corresponding to the PDF
       fields plus a mandatory '_output_filename' column. This Excel sheet
       is formatted as a Table for easier use.
    b) A text file (`_field_info.txt`) detailing expected values for
       non-text fields (like checkboxes or dropdowns).

2.  `fill-form`: Reads data from a populated Excel file (based on the generated
    template) and fills a copy of the PDF template for each data row. Each
    filled PDF is saved using the filename specified in the '_output_filename'
    column of the Excel sheet.

Features include colored logging for readability, a progress bar during form
filling, and an option to overwrite existing files in the output directory.

Dependencies:
    - pypdf
    - openpyxl
    - colorama
    - tqdm

Usage:
    python pybulkpdf.py generate-template --template <PDF_TEMPLATE> --output-dir <TEMPLATE_DIR>
    python pybulkpdf.py fill-form --template <PDF_TEMPLATE> --data-file <EXCEL_DATA> --output-dir <OUTPUT_DIR> [--overwrite]
"""

import argparse
import os
import sys
import logging
from typing import List, Tuple, Dict, Any, Optional, Set

# --- Third-Party Libraries ---
from pypdf import PdfReader, PdfWriter
from pypdf import errors as pypdf_errors
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import colorama
from colorama import Fore, Style, init
from tqdm import tqdm

# --- Constants ---
OUTPUT_FILENAME_COL: str = "_output_filename"
DEFAULT_SHEET_NAME: str = "Data"
DEFAULT_TABLE_NAME: str = "PDFData"
FIELD_INFO_SUFFIX: str = "_field_info.txt"
TEMPLATE_SUFFIX: str = "_template.xlsx"

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# --- Custom Colored Logging Formatter ---
class ColoredFormatter(logging.Formatter):
    """Custom logging formatter to add colors based on log level."""

    LEVEL_COLORS: Dict[int, str] = {
        logging.DEBUG: Fore.CYAN,
        logging.INFO: Fore.GREEN,
        logging.WARNING: Fore.YELLOW,
        logging.ERROR: Fore.RED,
        logging.CRITICAL: Fore.MAGENTA + Style.BRIGHT,
    }

    def format(self, record: logging.LogRecord) -> str:
        """Formats the log record with appropriate colors."""
        color = self.LEVEL_COLORS.get(record.levelno, Fore.WHITE)
        log_fmt = f"{color}{Style.BRIGHT}%(levelname)s{Style.RESET_ALL}: %(message)s"
        # Note: Timestamp could be added here if desired, e.g.,
        # log_fmt = f"%(asctime)s - {color}{Style.BRIGHT}%(levelname)s{Style.RESET_ALL}: %(message)s"
        # datefmt = '%Y-%m-%d %H:%M:%S' # Define this if using asctime
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)

# --- Logging Setup ---
def setup_logging() -> None:
    """Configures the root logger for console output with colors."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    # Clear existing handlers to prevent duplicate messages
    if logger.hasHandlers():
        logger.handlers.clear()
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(ColoredFormatter())
    logger.addHandler(console_handler)

# Call logging setup immediately
setup_logging()

# --- Helper Functions ---

def check_file_exists(filepath: str) -> None:
    """
    Checks if a file exists at the given path and is actually a file.
    Logs an error and exits the script if the check fails.

    Args:
        filepath: The path to the file to check.
    """
    if not os.path.exists(filepath):
        logging.error(f"Input file not found: {filepath}")
        sys.exit(1) # Fatal error
    if not os.path.isfile(filepath):
        logging.error(f"Input path is not a file: {filepath}")
        sys.exit(1) # Fatal error

def prepare_output_directory(dirpath: str, require_empty: bool = False, allow_overwrite: bool = False) -> None:
    """
    Checks and prepares the output directory. Creates it if non-existent.
    Optionally checks if it's empty based on parameters. Logs errors and exits on failure.

    Args:
        dirpath: The path to the output directory.
        require_empty: If True, checks if the directory is empty (unless allow_overwrite is True).
        allow_overwrite: If True, suppresses the "not empty" error when require_empty is True.
    """
    if os.path.exists(dirpath):
        # Path exists, check if it's a directory
        if not os.path.isdir(dirpath):
            logging.error(f"Output path '{dirpath}' exists but is not a directory.")
            sys.exit(1)
        # Path is a directory, check if it needs to be empty
        if require_empty and not allow_overwrite and os.listdir(dirpath):
            logging.error(f"Output directory '{dirpath}' is not empty. Use --overwrite flag or specify a different directory.")
            sys.exit(1)
        elif require_empty and allow_overwrite and os.listdir(dirpath):
            # Log warning if overwriting into a non-empty directory
            logging.warning(f"Output directory '{dirpath}' is not empty. Files may be overwritten.")
        # If directory exists, is valid, and passes emptiness check (if required), log usage.
        logging.info(f"Using existing output directory: {dirpath}")

    else:
        # Path does not exist, try to create it
        try:
            os.makedirs(dirpath)
            logging.info(f"Created output directory: {dirpath}")
        except (OSError, PermissionError) as e: # Broaden dir creation check
            logging.error(f"Error creating output directory '{dirpath}': {e}")
            sys.exit(1)

# --- Mode 1: Template Generation Logic ---

def generate_template_files(template_pdf_path: str, output_dir: str) -> None:
    """
    Analyzes a PDF template, extracts fillable field names, and generates an
    Excel template (.xlsx) formatted as a table, along with a text file
    containing information about non-text field types and expected values.

    Args:
        template_pdf_path: Path to the input PDF form template.
        output_dir: Directory where the generated template files will be saved.
    """
    logging.info(f"Starting template generation for: {template_pdf_path}")
    logging.info(f"Template files will be saved to: {output_dir}")

    check_file_exists(template_pdf_path)
    # For template generation, we don't require the directory to be empty.
    prepare_output_directory(output_dir, require_empty=False)

    try:
        # --- Read PDF Fields ---
        try:
            reader = PdfReader(template_pdf_path)
            # get_fields() can return None if no fields or AcroForm dict is missing
            fields: Optional[Dict[str, Any]] = reader.get_fields()
        except pypdf_errors.PdfReadError as e:
            logging.error(f"Error reading PDF template '{template_pdf_path}': {e}")
            sys.exit(1)
        except Exception as e:
             logging.error(f"Unexpected error opening or reading PDF '{template_pdf_path}': {e}")
             sys.exit(1)

        if not fields:
            logging.warning(f"No fillable form fields found in '{template_pdf_path}'. Template will only contain '{OUTPUT_FILENAME_COL}'.")
            field_names: List[str] = []
        else:
            field_names = list(fields.keys())
            logging.info(f"Found {len(field_names)} fields: {', '.join(field_names)}")

        # --- Generate XLSX Template ---
        base_filename = os.path.splitext(os.path.basename(template_pdf_path))[0]
        xlsx_filename = f"{base_filename}{TEMPLATE_SUFFIX}"
        xlsx_filepath = os.path.join(output_dir, xlsx_filename)
        # Ensure the mandatory output filename column is included
        xlsx_headers = field_names + [OUTPUT_FILENAME_COL]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = DEFAULT_SHEET_NAME
            ws.append(xlsx_headers) # Write header row

            # Create an Excel Table for better usability if headers exist
            if xlsx_headers:
                last_col_letter = get_column_letter(len(xlsx_headers))
                # Define table to cover header and first potential data row
                table_range = f"A1:{last_col_letter}2"
                tab = Table(displayName=DEFAULT_TABLE_NAME, ref=table_range)
                # Apply a standard style
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                logging.info(f"Formatted data range ({table_range}) as an Excel Table ('{DEFAULT_TABLE_NAME}').")
            else:
                 # This case should not happen if OUTPUT_FILENAME_COL is always added
                 logging.warning("Skipping Excel Table creation as no headers were generated.")

            wb.save(xlsx_filepath)
            logging.info(f"Generated Excel template: {xlsx_filepath}")

        except Exception as e:
            logging.error(f"Failed to create or save Excel template '{xlsx_filepath}': {e}")
            sys.exit(1) # Exit if template generation fails critically

        # --- Generate Field Info TXT (Only if fields were found) ---
        if fields:
            txt_filename = f"{base_filename}{FIELD_INFO_SUFFIX}"
            txt_filepath = os.path.join(output_dir, txt_filename)
            non_text_fields_info: List[str] = []

            for name, properties in fields.items():
                # Field properties dictionary structure can vary. Use .get for safety.
                field_type = properties.get('/FT') # Field Type (/Tx, /Btn, /Ch etc.)
                export_values: List[str] = []
                info: str = ""

                # Checkbox/Radio Button (/Btn)
                if field_type == '/Btn':
                    # Export values are often the keys in the Normal Appearance dictionary (/AP/N)
                    # Sometimes they might be in the field value itself (/V) if it's set.
                    # We prioritize /AP/N keys as they represent states.
                    ap_n_dict = properties.get('/AP', {}).get('/N', {})
                    if isinstance(ap_n_dict, dict): # It should be a dictionary-like object
                        export_values = list(ap_n_dict.keys())
                    # Remove '/Off' if present, as it's usually the default unselected state
                    if '/Off' in export_values:
                       export_values.remove('/Off')

                    info = f"Field '{name}' (Button): Expected values "
                    info += f"(e.g., {', '.join(export_values)})" if export_values else "(Check PDF for values like /Yes, /On)"
                    non_text_fields_info.append(info)

                # Choice Field (/Ch) - Dropdown/Listbox
                elif field_type == '/Ch':
                     # Options are usually in /Opt array [[display, export], ...] or [export, ...]
                     options = properties.get('/Opt', [])
                     info = f"Field '{name}' (Choice): Expected values "
                     if options:
                         # Check if options are [display, export] pairs or just export values
                         if isinstance(options[0], (list, tuple)) and len(options[0]) == 2:
                             export_values = [str(opt[1]) for opt in options] # Use the second element as export value
                         elif isinstance(options[0], (list, tuple)) and len(options[0]) == 1:
                             export_values = [str(opt[0]) for opt in options] # Use the first element if it's a single-item list/tuple
                         else: # Assume simple list of strings/values
                             export_values = [str(opt) for opt in options]
                     info += f": {', '.join(export_values)}" if export_values else "(Check PDF for options)"
                     non_text_fields_info.append(info)

                # Add checks for other field types if needed (e.g., /Sig for signature)

            # Write the info file if any non-text field info was gathered
            if non_text_fields_info:
                 try:
                     with open(txt_filepath, 'w', encoding='utf-8') as txtfile:
                         txtfile.write("Information about expected values for non-text PDF fields:\n")
                         txtfile.write("Note: Checkbox/Radio Button values are often '/Yes' or the specific value shown.\n")
                         txtfile.write("If unsure, test with a single row first.\n")
                         txtfile.write("=========================================================\n\n")
                         for line in non_text_fields_info:
                             txtfile.write(line + "\n")
                     logging.info(f"Generated field info file: {txt_filepath}")
                 except OSError as e:
                      # Log specific OS error but don't exit
                      logging.error(f"OS error writing field info file '{txt_filepath}': {e}")
                 except Exception as e:
                      # Log other errors but don't exit
                      logging.error(f"Failed to write field info file '{txt_filepath}': {e}")
            else:
                 logging.info("No specific non-text field information found to generate.")

        logging.info("Template generation completed successfully.")

    except Exception as e:
        # Catch-all for any other unexpected errors during the process
        logging.critical(f"An unexpected critical error occurred during template generation: {e}", exc_info=True)
        sys.exit(1)


# --- Mode 2: Form Filling Logic ---

def fill_pdf_forms(template_pdf_path: str, data_file_path: str, output_dir: str, overwrite: bool = False) -> None:
    """
    Fills PDF forms based on data read from an Excel file.

    Reads data row by row from the Excel sheet, uses the '_output_filename'
    column to name the output PDFs, and fills the fields found in both the
    PDF template and the Excel headers. Skips rows with errors and provides
    a summary at the end.

    Args:
        template_pdf_path: Path to the input PDF form template.
        data_file_path: Path to the input Excel (.xlsx) data file.
        output_dir: Directory where the filled PDF files will be saved.
        overwrite: If True, allows overwriting existing files in the output directory.
                   Defaults to False.
    """
    logging.info(f"Starting form filling using template: {template_pdf_path}")
    logging.info(f"Data source (Excel): {data_file_path}")
    logging.info(f"Filled PDFs will be saved to: {output_dir}")
    if overwrite:
        logging.warning("Overwrite flag set: Existing output files may be overwritten.")

    # --- Pre-Checks ---
    check_file_exists(template_pdf_path)
    check_file_exists(data_file_path)
    # Output directory preparation is handled in main() before calling this function

    failed_rows: List[Tuple[int, str]] = [] # Track rows that failed (row_num, reason)

    try:
        # --- Read PDF Fields (for comparison) ---
        try:
             pdf_reader_for_fields = PdfReader(template_pdf_path)
             pdf_fields: Optional[Dict[str, Any]] = pdf_reader_for_fields.get_fields()
             if not pdf_fields:
                 logging.error(f"No fillable fields found in template PDF: {template_pdf_path}")
                 sys.exit(1)
             pdf_field_names: Set[str] = set(pdf_fields.keys())
             logging.info(f"Template PDF fields found: {len(pdf_field_names)}")
        except pypdf_errors.PdfReadError as e:
             logging.error(f"Error reading PDF template fields from '{template_pdf_path}': {e}")
             sys.exit(1)
        except Exception as e:
             logging.error(f"Unexpected error reading PDF template fields '{template_pdf_path}': {e}")
             sys.exit(1)

        # --- Read Excel Data ---
        try:
            # data_only=True attempts to read cell values instead of formulas
            workbook = openpyxl.load_workbook(data_file_path, data_only=True)
            sheet = workbook.active # Use the active sheet
        except (InvalidFileException, FileNotFoundError) as excel_open_error:
             logging.error(f"Failed to open or find Excel file '{data_file_path}': {excel_open_error}")
             sys.exit(1)
        except Exception as e:
             logging.error(f"Error reading Excel file {data_file_path}: {e}")
             sys.exit(1)

        # --- Read and Validate Excel Headers ---
        header_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            # Read first row as headers
            xlsx_headers_raw = next(header_iter)
            # Clean headers (convert to string, strip whitespace, handle None)
            xlsx_headers: List[str] = [str(h).strip() if h is not None else '' for h in xlsx_headers_raw]
        except StopIteration:
            logging.error(f"Excel file appears empty or has no header row: {data_file_path}")
            sys.exit(1)

        if not xlsx_headers or all(h == '' for h in xlsx_headers):
            logging.error(f"Excel file has no valid headers in the first row: {data_file_path}")
            sys.exit(1)

        xlsx_header_set: Set[str] = set(xlsx_headers)

        # Check for the mandatory output filename column
        if OUTPUT_FILENAME_COL not in xlsx_header_set:
            logging.error(f"Required column '{OUTPUT_FILENAME_COL}' not found in Excel file headers: {data_file_path}")
            sys.exit(1)

        # --- Compare Headers and Determine Fields to Fill ---
        xlsx_data_fields: Set[str] = xlsx_header_set - {OUTPUT_FILENAME_COL}
        pdf_only_fields: Set[str] = pdf_field_names - xlsx_data_fields
        xlsx_only_fields: Set[str] = xlsx_data_fields - pdf_field_names

        if pdf_only_fields:
            logging.warning(f"PDF fields not found in Excel headers (will not be filled): {', '.join(sorted(list(pdf_only_fields)))}")
        if xlsx_only_fields:
            logging.warning(f"Excel headers not found in PDF fields (will be ignored): {', '.join(sorted(list(xlsx_only_fields)))}")

        common_fields: Set[str] = pdf_field_names.intersection(xlsx_data_fields)
        if not common_fields:
             logging.error("No common fields found between PDF template and Excel headers. Cannot proceed.")
             sys.exit(1)
        logging.info(f"Fields to be filled based on common headers: {len(common_fields)} ({', '.join(sorted(list(common_fields)))})")

        # --- Process Excel Rows ---
        # Estimate total rows for progress bar (may be inaccurate with trailing empty rows)
        total_rows = sheet.max_row - 1 if sheet.max_row > 1 else 0

        row_count = 0
        success_count = 0
        row_iterator = sheet.iter_rows(min_row=2, values_only=True) # Start from second row
        # Setup progress bar
        progress_bar = tqdm(
            row_iterator,
            total=total_rows,
            desc="Filling PDFs",
            unit="row",
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}{postfix}]",
            ncols=100 # Adjust width if needed
            )

        for row_index, row_values in enumerate(progress_bar):
            row_num = row_index + 2 # Excel row number is 1-based index + 1 for header

            # Skip rows where all cells are empty (common in Excel sheets)
            if all(v is None for v in row_values):
                continue

            row_count += 1
            current_output_filename: Optional[str] = None # For logging in case of error

            try:
                # Create dictionary mapping header names to cell values for this row
                row_dict = {header: value for header, value in zip(xlsx_headers, row_values)}

                # Get and validate the output filename for this row
                output_filename_raw = row_dict.get(OUTPUT_FILENAME_COL, '')
                current_output_filename = str(output_filename_raw).strip() if output_filename_raw is not None else ''

                if not current_output_filename:
                    logging.warning(f"Skipping row {row_num}: '{OUTPUT_FILENAME_COL}' is empty.")
                    failed_rows.append((row_num, f"'{OUTPUT_FILENAME_COL}' column is empty"))
                    continue # Skip to next row

                # Ensure filename ends with .pdf
                if not current_output_filename.lower().endswith('.pdf'):
                    current_output_filename += '.pdf'

                output_filepath = os.path.join(output_dir, current_output_filename)

                # Check if file exists and if overwriting is allowed
                if not overwrite and os.path.exists(output_filepath):
                     logging.warning(f"Skipping row {row_num}: Output file exists: {output_filepath} (use --overwrite to replace)")
                     failed_rows.append((row_num, f"Output file exists: {current_output_filename}"))
                     continue # Skip to next row

                # Update progress bar description for the current file
                progress_bar.set_postfix_str(f"Processing {current_output_filename}", refresh=True)

                # Prepare data dictionary for pypdf, ensuring values are strings
                # Note: pypdf might require specific values for checkboxes (e.g., '/Yes', '/Off')
                # Check the _field_info.txt or PDF for required values if simple strings don't work.
                fill_data = {
                    field: str(row_dict.get(field, '')) if row_dict.get(field) is not None else ''
                    for field in common_fields
                }

                # --- PDF Writing for the current row ---
                writer = None # Ensure writer is reset or defined
                try:
                    # Create a fresh writer by cloning the template inside the loop
                    writer = PdfWriter(clone_from=template_pdf_path)

                    # Iterate through all pages in the writer and update fields
                    # update_page_form_field_values only works per page
                    for page in writer.pages:
                        try:
                            # Attempt to update fields on the current page
                            writer.update_page_form_field_values(
                                page,
                                fields=fill_data
                            )
                        except KeyError:
                             # This can happen if a field in fill_data isn't on this specific page
                             # which is expected. We can safely ignore this.
                             pass
                        except Exception as page_update_error:
                            # Log if updating a specific page fails unexpectedly
                            page_num = writer.get_page_number(page)
                            logging.warning(f"Could not update fields on page {page_num+1} for {current_output_filename}: {page_update_error}")


                    # Remove /NeedAppearances flag if present (often helps compatibility)
                    # Check if AcroForm exists before trying to access it
                    if writer._root_object and "/AcroForm" in writer._root_object and "/NeedAppearances" in writer._root_object["/AcroForm"]:
                         writer._root_object["/AcroForm"].pop("/NeedAppearances")

                    # Write the filled PDF to the output file
                    with open(output_filepath, "wb") as output_stream:
                        writer.write(output_stream)

                    success_count += 1 # Increment success only if write completes

                except pypdf_errors.PdfReadError as pdf_read_err:
                    # Error reading the template during cloning
                    logging.error(f"Template PDF read error during cloning for row {row_num} ({current_output_filename}): {pdf_read_err}")
                    failed_rows.append((row_num, f"Template read error: {pdf_read_err}"))
                except FileNotFoundError as fnf_error:
                     # Error if output path is invalid during write
                     logging.error(f"File not found error during PDF write for row {row_num} ({current_output_filename}): {fnf_error}")
                     failed_rows.append((row_num, f"File path error: {fnf_error}"))
                except PermissionError as perm_error:
                     # Error if cannot write to output path
                     logging.error(f"Permission error during PDF write for row {row_num} ({current_output_filename}): {perm_error}")
                     failed_rows.append((row_num, f"File permission error: {perm_error}"))
                except Exception as pdf_write_error:
                    # Log other errors specific to PDF generation for this row
                    logging.error(f"PDF generation failed for row {row_num} ({current_output_filename}): {pdf_write_error}")
                    failed_rows.append((row_num, f"PDF write error: {pdf_write_error}"))
                finally:
                    # Ensure writer resources are potentially cleaned if applicable, though pypdf might handle this
                    # For safety, setting to None might help garbage collection if errors occurred mid-process
                    if writer:
                        writer = None # Allow garbage collection

            except KeyError as key_error:
                 # Error if expected column header (used in logic) is missing
                 logging.error(f"Data processing failed for row {row_num}: Missing expected column header {key_error}.")
                 failed_rows.append((row_num, f"Missing column {key_error}"))
            # Catch any other unexpected error while processing this row
            except Exception as row_error:
                # Catch any other unexpected error while processing this row
                logging.error(f"Unexpected error processing row {row_num}: {row_error}")
                failed_rows.append((row_num, f"Unexpected row error: {row_error}"))

    except Exception as e:
        # Catch-all for unexpected errors during setup or loop initialization
        logging.critical(f"An unexpected critical error occurred during the form filling process: {e}", exc_info=True)
        sys.exit(1)

    # --- Final Summary ---
    logging.info("-" * 40) # Separator for clarity
    logging.info("Form Filling Process Summary:")
    logging.info(f"Total data rows encountered: {row_count}")
    logging.info(f"Successfully generated PDFs: {success_count}")
    failures = len(failed_rows)
    if failures > 0:
        logging.warning(f"Rows with failures/skips: {failures}")
        # Log details for failed/skipped rows for review
        for row_num, reason in failed_rows:
             logging.warning(f"  - Row {row_num}: {reason}")
    else:
        logging.info("All processed rows generated PDFs successfully.")
    logging.info("-" * 40)


# --- Main Execution Logic ---
def main() -> None:
    """
    Parses command-line arguments, prepares the environment, and calls the
    appropriate function based on the selected mode ('generate-template' or 'fill-form').
    """
    parser = argparse.ArgumentParser(
        description="PyBulkPDF: Fill PDF forms from Excel (.xlsx) data.",
        formatter_class=argparse.RawTextHelpFormatter # Preserves formatting in help text
        )

    # Define subparsers for the two modes
    subparsers = parser.add_subparsers(
        dest='mode',
        required=True,
        help='Choose the operation mode: "generate-template" or "fill-form"'
        )

    # --- Subparser for Template Generation ---
    parser_gen = subparsers.add_parser(
        'generate-template',
        help='Analyze a PDF form and generate template files (XLSX Table and TXT).'
        )
    parser_gen.add_argument(
        '--template', '-t',
        required=True,
        help='Path to the input PDF form template file.'
        )
    parser_gen.add_argument(
        '--output-dir', '-o',
        required=True,
        help='Directory to save the generated template XLSX and field info TXT files.'
        )
    parser_gen.set_defaults(func=generate_template_files) # Link to function

    # --- Subparser for Form Filling ---
    parser_fill = subparsers.add_parser(
        'fill-form',
        help='Fill PDF forms using data from an Excel (.xlsx) file.'
        )
    parser_fill.add_argument(
        '--template', '-t',
        required=True,
        help='Path to the input PDF form template file.'
        )
    parser_fill.add_argument(
        '--data-file', '-d',
        required=True,
        help='Path to the input Excel (.xlsx) data file (based on generated template).'
        )
    parser_fill.add_argument(
        '--output-dir', '-o',
        required=True,
        help='Directory to save the filled PDF output files.'
        )
    parser_fill.add_argument(
        '--overwrite',
        action='store_true', # Flag, doesn't take a value
        help='Allow overwriting existing files in the output directory. Use with caution.'
        )
    parser_fill.set_defaults(func=fill_pdf_forms) # Link to function

    # --- Parse Arguments ---
    args = parser.parse_args()

    # --- Prepare Output Directory and Execute Selected Mode ---
    output_dir_to_check = args.output_dir

    try:
        if args.mode == 'generate-template':
            # Template generation allows outputting to an existing (non-empty) directory
            prepare_output_directory(output_dir_to_check, require_empty=False)
            # Call the function with relevant args
            args.func(args.template, args.output_dir)

        elif args.mode == 'fill-form':
            # Form filling requires the directory to be empty unless --overwrite is specified
            prepare_output_directory(output_dir_to_check, require_empty=True, allow_overwrite=args.overwrite)
            # Call the function with relevant args, including overwrite status
            args.func(args.template, args.data_file, args.output_dir, args.overwrite)

    except SystemExit:
         # Raised by sys.exit() in helper functions on fatal errors
         logging.critical("Execution halted due to fatal error.")
         # Optionally exit with a specific code: sys.exit(1)
    except Exception as e:
         # Catch any other unexpected errors in the main execution flow
         logging.critical(f"An unexpected error occurred in main execution: {e}", exc_info=True)
         sys.exit(1)


if __name__ == "__main__":
    # This ensures the main function runs only when the script is executed directly
    main()
